import openpyxl

#Instalar "pip install openpyxl" (es una libreria)

class FunExcel():

    #Inicializa la clase 
    def __init__(self,driver):
        self.driver = driver

    def getRowCount(file, path, sheetName):
        Worbook=openpyxl.load_workbook(path)
        sheet=Worbook[sheetName]
        return (sheet.max_row)
    
    def getColumnCount(file, sheetName):
        Worbook=openpyxl.load_workbook(file)
        sheet=Worbook[sheetName]
        return (sheet.max_column)

    def redData(file, path, sheetName, rownum, columna):
        Worbook=openpyxl.load_workbook(path)
        sheet=Worbook[sheetName]
        #.cell = LEE. 
        return sheet.cell(row=rownum, column=columna).value
    
    def writeData(file, path, sheetName, rownum, columna, data):
        Worbook=openpyxl.load_workbook(path)
        sheet=Worbook[sheetName] 
        sheet.cell(row=rownum, column=columna).value = data
        Worbook.save(path)


